import openpyxl

file_path = r'C:\Users\Geo_C\OneDrive\Documents\Rendimiento de Coextruido - EMSULA.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"--- Sheet: {sheet_name} ---")
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    
    # Read the first data row (row 2 or 3 usually) to see formulas
    for row_idx in range(2, min(5, sheet.max_row + 1)):
        row = sheet[row_idx]
        formulas = []
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                formulas.append(val)
            else:
                formulas.append("VAL")
        if any(f != "VAL" for f in formulas):
            print(f"Row {row_idx} formulas: {formulas}")
            break
    print("\n")
